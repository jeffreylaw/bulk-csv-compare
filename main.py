import csv
import sys
import os
import glob
import json
import re
import platform
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import tkinter
from tkinter import messagebox


def get_key(dic,value):
    for i in dic:
        if dic[i] == value:
            return i

def compare_files(file1, file2, wb):
    filename = os.path.basename(file1).split('-')[0]
    with open(file1, 'r', encoding='utf8') as file1, open(file2, 'r', encoding='utf8') as file2:
        reader1 = csv.reader(file1)
        reader2 = csv.reader(file2)

        file1_participants = ['Me']
        file2_participants = ['Me']

        for i in range(7):
            if i == 1:
                downloader_participant = next(reader1)[0].split(': ')[1].split('@')[0]
                file1_participants.append(downloader_participant)

                downloader_participant = next(reader2)[0].split(': ')[1].split('@')[0]
                file2_participants.append(downloader_participant)
                continue

            if i == 3:
                participants = next(reader1)[0]
                if len(participants) > 20:
                    participants = participants.split()[1].split(';')[:-1]
                    participants = [i.split('@')[0] for i in participants]
                    file1_participants = file1_participants + participants

                participants = next(reader2)[0]
                if len(participants) > 20:
                    participants = participants.split()[1].split(';')[:-1]
                    participants = [i.split('@')[0] for i in participants]
                    file2_participants = file2_participants + participants
                continue

            next(reader1)
            next(reader2)


        file1_chat = {}
        file2_chat = {}

        participants_regex = [ '^' + i + r'\d{4}-\d{2}-\d{2}' for i in file1_participants]
        combined = "(" + ")|(".join(participants_regex) + ")"

        message_num = 1
        for i, l, in enumerate(reader1, start=1):
            current_line = ''.join(l)
            if re.match(combined, current_line):
                file1_chat[message_num] = current_line
                message_num += 1
            else:
                if not current_line:
                    file1_chat[message_num-1] = file1_chat[message_num-1] + '\n\n'
                else:
                    file1_chat[message_num-1] = file1_chat[message_num-1] + current_line

        message_num = 1
        for i, l, in enumerate(reader2, start=1):
            current_line = ''.join(l)
            if re.match(combined, current_line):
                file2_chat[message_num] = current_line
                message_num += 1
            else:
                if not current_line:
                    file2_chat[message_num-1] = file2_chat[message_num-1] + '\n\n'
                else:
                    file2_chat[message_num-1] = file2_chat[message_num-1] + current_line

        result_dict = {filename: 'Pass'}

        for i in file2_chat.values():
            if i in file1_chat.values():
                key = get_key(file1_chat, i)
                del file1_chat[key]

        if len(file1_chat) > 0:
            result_dict = {filename: 'Fail'}
            heading_font = Font(bold=True)

            ws = wb.create_sheet(title=filename)
            ws.column_dimensions['A'].width = 20
            ws['A1'] = 'Chatroom name:'
            ws['B1'] = filename

            ws['A3'].font = heading_font
            ws['B3'].font = heading_font

            ws['A3'] = 'Message number:'
            ws['B3'] = 'Contents:'

            current_row = 4
            for key, value in file1_chat.items():
                ws['A' + str(current_row)] = key
                ws['B' + str(current_row)] = value
                current_row += 1
    return result_dict


if __name__ == '__main__':
    results = {}
    if len(sys.argv) == 1:
        if not os.path.exists('before') or not os.path.isdir('before'):
            print('Missing \'before\' directory')
            exit(1)
        if not os.path.exists('after') or not os.path.isdir('after'):
            print('Missing \'after\' directory')
            exit(1)

        csv_files = {}

        if platform.system() == 'Darwin':
            before_path = os.getcwd() + '/before'
            after_path = os.getcwd() + '/after'
        else:
            before_path = os.getcwd() + '\\before'
            after_path = os.getcwd() + '\\after'

        for root, dirs, files in os.walk(before_path):
            dirs.clear()
            for file in files:
                if file.endswith('.csv'):
                    csv_files[file] = 1

        for root, dirs, files in os.walk(after_path):
            dirs.clear()
            for file in files:
                if file.endswith('.csv'):
                    if file in csv_files.keys():
                        csv_files[file] = csv_files[file] + 1
                    else:
                        csv_files[file] = 1

        wb = openpyxl.Workbook()

        sheet = wb.active
        sheet.title = 'Results'


        for i in csv_files:
            if csv_files[i] == 2:
                if platform.system() == 'Darwin':
                    result = compare_files(
                        before_path + '/' + i, after_path + '/' + i, wb)            
                else:
                    result = compare_files(
                        before_path + '\\' + i, after_path + '\\' + i, wb)
                results.update(result)
            elif csv_files[i] == 1:
                results[i] = 'Failed (Missing second file to compare with)'
            else:
                results[i] = 'Failed (Somehow found more than 2 copies)'


    ws = wb['Results']
    ws.column_dimensions['A'].width = 30

    pass_font = Font(color='008000')
    fail_font = Font(color='FF0000')

    current_row = 5
    for key, value in results.items():
        ws['A' + str(current_row)] = key
        ws['B' + str(current_row)] = value
        if value == 'Pass':
            ws['B' + str(current_row)].font = pass_font
        else:
            ws['B' + str(current_row)].font = fail_font
        current_row += 1

    now = datetime.now()

    ws['A1'] = 'Created:'
    ws['B1'] = now.strftime('%d-%b-%Y %I:%M:%S%p')

    heading_font = Font(bold=True)
    ws['A4'].font = heading_font
    ws['A4'] = 'Name:'
    ws['B4'].font = heading_font
    ws['B4'] = 'Synced:'
    ws.freeze_panes = 'A5'
    wb.save('results_' + now.strftime('%d.%b.%Y_%I.%M.%S%p') + '.xlsx')
